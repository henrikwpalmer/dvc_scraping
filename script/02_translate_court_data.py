"""
translate_court_data.py
-----------------------
Reads the combined ALL_COURTS workbook produced by sudrf_scraper.py
(cases_output_ALL_COURTS.xlsx), translates the Russian free-text columns
to English using deep-translator (free, no API key required), and writes
the results to a second "ENGLISH" tab in the same workbook -- the original
sheet (with the untranslated Russian data) is left untouched.

Optimisations:
  - Batches unique strings so each unique value is translated only once,
    regardless of how many rows it appears in.
  - Skips fields that don't need translation (dates, case numbers, URLs,
    row numbers, court names already transliterated, etc.).
  - Falls back to the original text if a translation call fails.
  - Prints a progress summary so you can track long runs.

Usage:
    pip install deep-translator openpyxl
    python translate_court_data.py

    # Or specify a different file:
    python translate_court_data.py my_data.xlsx
"""

import sys
import re
import time
import random

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from deep_translator import GoogleTranslator

# ── Configuration ──────────────────────────────────────────────────────────────

INPUT_FILE  = sys.argv[1] if len(sys.argv) > 1 else "../output/cases_output_ALL_COURTS_kherson.xlsx"
SOURCE_LANG = "ru"
TARGET_LANG = "en"

# pandas' DataFrame.to_excel() (used by sudrf_scraper.py) writes to a sheet
# named "Sheet1" by default. If that's not found, we fall back to whichever
# sheet is first in the workbook, so this still works if you've renamed it.
SOURCE_SHEET = "Sheet1"
OUTPUT_SHEET = "ENGLISH"

# Columns to translate, matched against the exact header text written by
# sudrf_scraper.py. Only "judge" and "solution" carry free-form Russian
# prose in this workbook -- everything else is a date, URL, case number,
# page number, or already-Latin metadata (court_name, notes, source_base_url).
# Add more column names here (matching your header row exactly) if you add
# fields to the scraper later, e.g. if a "category" column gets added.
TRANSLATE_COLUMNS = {
    "judge",
    "solution",
}

# Polite delay range between API calls (seconds)
DELAY = (0.15, 0.35)

# Max characters per translate() call (Google's limit is ~5000)
MAX_CHARS = 4500


# ── Styling ────────────────────────────────────────────────────────────────────

HDR_FILL   = PatternFill("solid", start_color="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT  = Font(name="Arial", size=10)
ALT_FILL   = PatternFill("solid", start_color="EBF3FB")
THIN = Border(
    left=Side(style="thin"),  right=Side(style="thin"),
    top=Side(style="thin"),   bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


# ── Translation ────────────────────────────────────────────────────────────────

def needs_translation(text: str) -> bool:
    """Skip blanks, dates, URLs, pure numbers, and text with no Cyrillic."""
    if not text or not text.strip():
        return False
    t = text.strip()
    if t.startswith("http"):
        return False
    if re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", t):
        return False
    if re.fullmatch(r"\d{2}:\d{2}", t):
        return False
    if re.fullmatch(r"[\d\-]+", t):
        return False
    # If there's no Cyrillic in the string at all, there's nothing to
    # translate (e.g. an already-transliterated court name, or blank notes).
    if not re.search(r"[А-Яа-яЁё]", t):
        return False
    return True


def build_translation_cache(unique_strings: set) -> dict:
    """
    Translate all unique strings in one pass.
    Returns {original: translated} dict.
    """
    translator = GoogleTranslator(source=SOURCE_LANG, target=TARGET_LANG)
    cache = {}
    to_translate = [s for s in unique_strings if needs_translation(s)]

    print(f"\nTranslating {len(to_translate)} unique strings "
          f"(skipping {len(unique_strings) - len(to_translate)} non-text values)…")

    for i, text in enumerate(to_translate, 1):
        chunk = text[:MAX_CHARS]
        try:
            translated = translator.translate(chunk)
            cache[text] = translated if translated else text
        except Exception as exc:
            print(f"  [WARN] Failed to translate string #{i}: {exc!r}")
            cache[text] = text  # fall back to original

        if i % 25 == 0 or i == len(to_translate):
            print(f"  {i}/{len(to_translate)} translated…")

        time.sleep(random.uniform(*DELAY))

    for s in unique_strings:
        if s not in cache:
            cache[s] = s

    return cache


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    print(f"Loading: {INPUT_FILE}")
    wb = load_workbook(INPUT_FILE)

    source_sheet = SOURCE_SHEET
    if source_sheet not in wb.sheetnames:
        source_sheet = wb.sheetnames[0]
        print(f"[i] Sheet '{SOURCE_SHEET}' not found -- using first sheet "
              f"'{source_sheet}' instead. Available sheets: {wb.sheetnames}")

    ws_src = wb[source_sheet]

    # Read header row → map column index to label
    headers = {}
    for cell in ws_src[1]:
        if cell.value:
            headers[cell.column] = str(cell.value)

    translate_col_indices = {
        col_idx for col_idx, label in headers.items()
        if label in TRANSLATE_COLUMNS
    }

    if not translate_col_indices:
        print(f"[!] None of {sorted(TRANSLATE_COLUMNS)} were found in the "
              f"header row ({list(headers.values())}). Nothing will be "
              f"translated -- the ENGLISH tab will just be a copy. Check "
              f"that TRANSLATE_COLUMNS matches your actual column headers.")
    else:
        print(f"Columns to translate: "
              f"{[headers[i] for i in sorted(translate_col_indices)]}")

    # ── Collect all unique strings that need translation ───────────────────────
    unique_strings = set()
    data_rows = list(ws_src.iter_rows(min_row=2, values_only=True))

    for row in data_rows:
        for col_idx in translate_col_indices:
            val = row[col_idx - 1]  # iter_rows is 0-indexed per row tuple
            if val and str(val).strip():
                unique_strings.add(str(val))

    print(f"Unique non-empty values in translatable columns: {len(unique_strings)}")

    # ── Translate ──────────────────────────────────────────────────────────────
    cache = build_translation_cache(unique_strings)

    # ── Remove existing ENGLISH sheet if re-running ───────────────────────────
    if OUTPUT_SHEET in wb.sheetnames:
        del wb[OUTPUT_SHEET]

    # ── Create ENGLISH sheet ───────────────────────────────────────────────────
    ws_out = wb.create_sheet(OUTPUT_SHEET)

    # Header row
    for cell in ws_src[1]:
        new_cell = ws_out.cell(row=1, column=cell.column, value=cell.value)
        new_cell.fill      = HDR_FILL
        new_cell.font      = HDR_FONT
        new_cell.border    = THIN
        new_cell.alignment = CENTER

    ws_out.freeze_panes = "C2"  # keep case_number/case_link visible while scrolling

    # Data rows
    for row_idx, row in enumerate(data_rows, 2):
        is_alt = row_idx % 2 == 0
        for col_idx, label in headers.items():
            original = row[col_idx - 1]
            val = str(original) if original is not None else ""

            if col_idx in translate_col_indices and val in cache:
                val = cache[val]

            cell = ws_out.cell(row=row_idx, column=col_idx, value=val)
            cell.font      = DATA_FONT
            cell.border    = THIN
            cell.alignment = LEFT
            if is_alt:
                cell.fill = ALT_FILL

    # Column widths
    for col_idx, label in headers.items():
        col_letter = get_column_letter(col_idx)
        max_len = max(
            len(str(ws_out.cell(row=r, column=col_idx).value or ""))
            for r in range(1, len(data_rows) + 2)
        )
        ws_out.column_dimensions[col_letter].width = min(max(max_len + 2, 12), 55)

    ws_out.auto_filter.ref = ws_out.dimensions

    # ── Save ───────────────────────────────────────────────────────────────────
    wb.save(INPUT_FILE)
    print(f"\n✓ Done — '{OUTPUT_SHEET}' tab written to {INPUT_FILE}")
    print(f"  {len(data_rows)} rows translated across "
          f"{len(translate_col_indices)} columns.")


if __name__ == "__main__":
    main()

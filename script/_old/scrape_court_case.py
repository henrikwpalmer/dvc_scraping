"""
scrape_court_case.py
--------------------
Scrapes case proceedings from Russian sudrf.ru court portals and writes
results to a flat, case-event-level Excel sheet — one row per event,
with case-level fields repeated on every row.

Usage (single case):
    python scrape_court_case.py

Usage (multiple cases):
    Add URLs to CASE_URLS list at the top.

Dependencies:
    pip install requests beautifulsoup4 openpyxl

Anti-blocking:
  1. Realistic Chrome User-Agent + Accept / Accept-Language / Referer headers
  2. Session warm-up on court homepage to obtain a session cookie
  3. Random jitter between requests and retries
  4. Windows-1251 encoding via response.apparent_encoding
"""

import re
import time
import random
import requests
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Configuration ──────────────────────────────────────────────────────────────

CASE_URLS = [
    (
        "https://amv--dnr.sudrf.ru/modules.php"
        "?name=sud_delo&srv_num=1&name_op=case"
        "&case_id=24345375"
        "&case_uid=11312572-a74d-441c-becc-db28c1783dc0"
        "&delo_id=1540005&new="
    ),
    # Add more URLs here:
    # "https://amv--dnr.sudrf.ru/modules.php?name=sud_delo&...",
]

OUTPUT_FILE = "court_case_data.xlsx"

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "ru-RU,ru;q=0.9,en-US;q=0.8,en;q=0.7",
    "Accept-Encoding": "gzip, deflate, br",
    "Connection": "keep-alive",
}

MAX_RETRIES = 3
RETRY_DELAY = (5, 10)


# ── Fetch ──────────────────────────────────────────────────────────────────────

def build_session(base_url: str) -> requests.Session:
    session = requests.Session()
    session.headers.update(HEADERS)
    try:
        r = session.get(base_url, timeout=15)
        r.raise_for_status()
        print(f"  Warm-up OK ({r.status_code})")
    except requests.RequestException as e:
        print(f"  Warm-up failed (non-fatal): {e}")
    time.sleep(random.uniform(1.0, 2.5))
    return session


def fetch_page(url: str) -> BeautifulSoup:
    base = "/".join(url.split("/")[:3]) + "/"
    session = build_session(base)
    for attempt in range(1, MAX_RETRIES + 1):
        try:
            resp = session.get(url, timeout=20, headers={"Referer": base})
            resp.raise_for_status()
            resp.encoding = resp.apparent_encoding or "windows-1251"
            print(f"  Fetched OK (attempt {attempt}, encoding={resp.encoding})")
            return BeautifulSoup(resp.text, "html.parser")
        except requests.RequestException as exc:
            print(f"  Attempt {attempt}/{MAX_RETRIES} failed: {exc}")
            if attempt < MAX_RETRIES:
                delay = random.uniform(*RETRY_DELAY)
                print(f"  Retrying in {delay:.1f}s…")
                time.sleep(delay)
            else:
                raise


# ── Parse ──────────────────────────────────────────────────────────────────────

def clean(text) -> str:
    if text is None:
        return ""
    return " ".join(str(text).split()).strip()


def is_date(text: str) -> bool:
    return bool(re.fullmatch(r"\d{2}\.\d{2}\.\d{4}", text.strip()))


def parse_case_info(soup: BeautifulSoup) -> dict:
    """
    Extract the case metadata block. On sudrf.ru this is a key-value table
    where each row has a <b>Label</b> cell followed by a value cell.

    Known real labels from the live site:
      Уникальный идентификатор дела  → judicial_uid
      Дата поступления               → date_filed
      Категория дела                 → case_category
      Судья                          → judge
      Признак рассмотрения дела      → hearing_type
    The case number lives in an <h2> or similar heading.
    """
    KEY_MAP = {
        "уникальный идентификатор дела":  "judicial_uid",
        "уникальный идентификатор":       "judicial_uid",
        "дата поступления":               "date_filed",
        "категория дела":                 "case_category",
        "судья":                          "judge",
        "признак рассмотрения дела":      "hearing_type",
        "текущий статус":                 "current_stage",
        "результат":                      "current_stage",
    }

    info = {}

    # Case number: find the text node containing "Дело №" / "Дело N" pattern.
    # It may be in a heading or a plain element, not necessarily <h2>.
    case_no_re = re.compile(r"Дело\s*[№N]\s*[\w\-/]+.*", re.IGNORECASE)
    candidate = soup.find(string=case_no_re)
    if candidate:
        m = case_no_re.search(clean(candidate))
        if m:
            info["case_number"] = clean(m.group(0))
    if "case_number" not in info:
        # Fallback: any heading that looks like it contains a case number
        for tag in ("h1", "h2", "h3", "h4", "h5"):
            for h in soup.find_all(tag):
                txt = clean(h.get_text())
                if re.search(r"\d+-\d+/\d{4}", txt):
                    info["case_number"] = txt
                    break
            if "case_number" in info:
                break

    # Key-value rows: a cell containing <b> or <strong> is the label
    for table in soup.find_all("table"):
        for row in table.find_all("tr"):
            cells = row.find_all(["td", "th"])
            if len(cells) < 2:
                continue
            bold = cells[0].find("b") or cells[0].find("strong")
            if not bold:
                continue
            label_raw = clean(bold.get_text())
            value = clean(cells[1].get_text())
            if not label_raw or not value:
                continue
            canon = KEY_MAP.get(label_raw.lower(), label_raw)
            info[canon] = value

    return info


def parse_proceedings(soup: BeautifulSoup) -> list:
    """
    Extract the proceedings/events table.

    Real column headers on the live site (windows-1251 decoded):
      Наименование события            → event_name
      Дата                            → event_date
      Время                           → event_time
      Результат заседания             → event_result  (may be blank)
      Дата публикации / опубликования → event_published_date

    The table is identified by having BOTH a date-column header AND
    an event-name column header.
    """
    def classify(h: str):
        """Map a header cell to a canonical field name (order matters)."""
        if "наименование события" in h or "мероприятие" in h or h == "событие":
            return "event_name"
        if "дата публикации" in h or "дата опубликован" in h or "опубликован" in h:
            return "event_published_date"
        if "результат" in h or "итог" in h:
            return "event_result"
        if "основания" in h:  # "Основания для..." column on some courts
            return "event_basis"
        if "время" in h:
            return "event_time"
        if "дата" in h:
            return "event_date"
        return None

    for table in soup.find_all("table"):
        header_row = table.find("tr")
        if not header_row:
            continue
        raw_headers = [clean(c.get_text()) for c in header_row.find_all(["th", "td"])]
        lower_headers = [h.lower() for h in raw_headers]

        col_map = {}
        for i, h in enumerate(lower_headers):
            field = classify(h)
            if field:
                col_map[i] = field

        # Qualify only if this is the events table (has event_name AND a date)
        fields = set(col_map.values())
        if "event_name" not in fields or "event_date" not in fields:
            continue

        rows = []
        for tr in table.find_all("tr")[1:]:
            cells = tr.find_all(["td", "th"])
            if not cells:
                continue
            row = {col_map[i]: clean(cells[i].get_text())
                   for i in range(len(cells)) if i in col_map}
            # Keep a row only if it has an actual event name (skips stray header/blank rows)
            if row.get("event_name", "").strip():
                rows.append(row)

        if rows:
            return rows

    return []


def parse_participants(soup: BeautifulSoup) -> list:
    """
    Extract the participants table.

    Real column headers on the live site:
      Вид лица, участвующего в деле   → role
      Фамилия / наименование          → name
      (sometimes ИНН, Адрес, etc.)
    """
    def classify(h: str):
        # role column: "Вид лица, участвующего в деле"
        if "вид лица" in h or "роль" in h or "тип участника" in h:
            return "role"
        # name column: "Фамилия / наименование" — but NOT "Наименование события"
        if ("фамилия" in h or "фио" in h or "ф.и.о" in h
                or ("наименование" in h and "событи" not in h)):
            return "name"
        if "инн" in h:
            return "inn"
        return None

    for table in soup.find_all("table"):
        header_row = table.find("tr")
        if not header_row:
            continue
        raw_headers = [clean(c.get_text()) for c in header_row.find_all(["th", "td"])]
        lower_headers = [h.lower() for h in raw_headers]

        col_map = {}
        for i, h in enumerate(lower_headers):
            field = classify(h)
            if field:
                col_map[i] = field

        # The participants table must have a role column (the events table won't)
        if "role" not in set(col_map.values()):
            continue

        rows = []
        for tr in table.find_all("tr")[1:]:
            cells = tr.find_all(["td", "th"])
            if not cells:
                continue
            row = {col_map[i]: clean(cells[i].get_text())
                   for i in range(len(cells)) if i in col_map}
            if any(row.values()):
                rows.append(row)

        if rows:
            return rows

    return []


def participants_to_columns(participants: list) -> dict:
    """
    Collapse participants into three flat strings: plaintiffs, defendants, other.
    """
    PLAINTIFF_KEYS  = {"истец", "заявитель", "plaintiff"}
    DEFENDANT_KEYS  = {"ответчик", "defendant"}

    plaintiffs, defendants, others = [], [], []
    for p in participants:
        role = p.get("role", "").lower()
        name = p.get("name", "") or p.get("role", "")
        if any(k in role for k in PLAINTIFF_KEYS):
            plaintiffs.append(name)
        elif any(k in role for k in DEFENDANT_KEYS):
            defendants.append(name)
        else:
            label = p.get("role", "")
            entry = f"{label}: {name}" if label and name and label != name else (name or label)
            if entry:
                others.append(entry)

    return {
        "participants_plaintiffs":  "; ".join(plaintiffs),
        "participants_defendants":  "; ".join(defendants),
        "participants_other":       "; ".join(others),
    }


def scrape_case(url: str) -> list:
    """Scrape one case URL → list of flat event-row dicts."""
    print(f"\nScraping: {url}")
    soup = fetch_page(url)

    case_info    = parse_case_info(soup)
    proceedings  = parse_proceedings(soup)
    participants = parse_participants(soup)
    part_cols    = participants_to_columns(participants)

    print(f"  case_info fields : {list(case_info.keys())}")
    print(f"  proceedings rows : {len(proceedings)}")
    print(f"  participant rows : {len(participants)}")

    case_fields = {"source_url": url, **case_info, **part_cols}

    if not proceedings:
        return [{**case_fields,
                 "event_name": "", "event_date": "", "event_time": "",
                 "event_result": "", "event_published_date": ""}]

    return [{**case_fields, **event} for event in proceedings]


# ── Excel output ──────────────────────────────────────────────────────────────

# Ordered, canonical output columns
COLUMN_ORDER = [
    "source_url",
    "case_number",
    "judicial_uid",
    "date_filed",
    "hearing_type",
    "case_category",
    "judge",
    "current_stage",
    "participants_plaintiffs",
    "participants_defendants",
    "participants_other",
    "event_name",
    "event_date",
    "event_time",
    "event_result",
    "event_basis",
    "event_published_date",
]

COLUMN_LABELS = {
    "source_url":               "Source URL",
    "case_number":              "Case Number",
    "judicial_uid":             "Judicial UID",
    "date_filed":               "Date Filed",
    "hearing_type":             "Hearing Type",
    "case_category":            "Case Category",
    "judge":                    "Judge",
    "current_stage":            "Current Stage",
    "participants_plaintiffs":  "Plaintiff(s)",
    "participants_defendants":  "Defendant(s)",
    "participants_other":       "Other Participants",
    "event_name":               "Event",
    "event_date":               "Event Date",
    "event_time":               "Event Time",
    "event_result":             "Event Result",
    "event_basis":              "Basis / Grounds",
    "event_published_date":     "Published Date",
}

EVENT_COLS = {"event_name","event_date","event_time","event_result","event_basis","event_published_date"}

HDR_FILL  = PatternFill("solid", start_color="1F4E79")
HDR_FONT  = Font(bold=True, color="FFFFFF", name="Arial", size=10)
DATA_FONT = Font(name="Arial", size=10)
CASE_FILL  = PatternFill("solid", start_color="EBF3FB")
EVENT_FILL = PatternFill("solid", start_color="FFF9F0")
THIN = Border(left=Side(style="thin"),  right=Side(style="thin"),
              top=Side(style="thin"),   bottom=Side(style="thin"))
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)


def write_excel(all_rows: list, filepath: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Case Events"

    # Full column list: predefined order + any extra columns from the scrape
    seen = set(COLUMN_ORDER)
    extra_cols = [k for row in all_rows for k in row if k not in seen and not seen.add(k)]
    full_cols = COLUMN_ORDER + extra_cols

    # Header
    ws.append([COLUMN_LABELS.get(c, c) for c in full_cols])
    for ci in range(1, len(full_cols) + 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill = HDR_FILL; cell.font = HDR_FONT
        cell.border = THIN; cell.alignment = CENTER

    ws.freeze_panes = "C2"

    # Data
    for ri, row_data in enumerate(all_rows, 2):
        is_alt = ri % 2 == 0
        for ci, col_key in enumerate(full_cols, 1):
            cell = ws.cell(row=ri, column=ci, value=row_data.get(col_key, ""))
            cell.font = DATA_FONT; cell.border = THIN; cell.alignment = LEFT
            if is_alt:
                cell.fill = EVENT_FILL if col_key in EVENT_COLS else CASE_FILL

    # Column widths
    for ci, col_key in enumerate(full_cols, 1):
        max_len = max(
            len(str(ws.cell(row=r, column=ci).value or ""))
            for r in range(1, len(all_rows) + 2)
        )
        ws.column_dimensions[get_column_letter(ci)].width = min(max(max_len + 2, 12), 55)

    ws.auto_filter.ref = ws.dimensions
    wb.save(filepath)
    print(f"\n✓ Saved → {filepath}  ({len(all_rows)} event rows)")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    all_rows = []
    for url in CASE_URLS:
        try:
            rows = scrape_case(url)
            all_rows.extend(rows)
            print(f"  → {len(rows)} event row(s) collected")
            if url != CASE_URLS[-1]:
                time.sleep(random.uniform(2, 4))
        except Exception as e:
            print(f"  ERROR scraping {url}: {e}")
            all_rows.append({"source_url": url, "case_number": "ERROR", "event_name": str(e)})

    if all_rows:
        write_excel(all_rows, "../output/" + OUTPUT_FILE)
    else:
        print("No data collected.")


if __name__ == "__main__":
    main()
